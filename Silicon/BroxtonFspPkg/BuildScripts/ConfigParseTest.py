##
# 
# @package ConfigParseTest
# This is a standalone tool that will run an array of test Builds for
# different configurations. It will create an excel file containing the test configurations as well the result.
# This is not a part of the build process, but a validation tool.
# 

import ConfigParser
from openpyxl import Workbook
from openpyxl.styles import Color, Fill, colors
import openpyxl
import subprocess

# OptionToColumn = {
#                  'SiliconType':  2,
#                  'BuildTarget':  3,
#                  'NoFsp':        4,
#                  'FspWrapper':   5,
#                  'Quiet':        6,
#                  'Logging':      7,
#                  'CleanBuild':   8,
#                  'WinDb':        9,
#                  'SvBios':       10,
#                  'PpvBios':      11,
#                  'CompactSle':   12,
#                  'Win7':         13,
#                  'NoGopUpdate':  14,
#                  'UpdatePmc':    15,
#                  }

OptionToColumn = {
                 'BuildTarget':  7,
                 'SiliconType':  6,
                 'NoMrc':         5,
                 'NoMrcDbg':        4,
                 'CompactSle':   3,
                 'PpvBios':        2,
                 'SvBios':        1,
                 }

NonBools = {
            'SiliconType': {'0': 'Virtual', '1': 'Real'},
            'BuildTarget': {'0': 'Debug', '1': 'Release'},
            # 'Arch'       : {'0': 'IA32', '1': 'X64'}
           }

BaseBinary = int(128)

def ReverseNumbering(Number, MaxNumber):
    return (MaxNumber + 1) - Number

def InitializeWorkSheet(WorkSheet):

    TopRow = 1

    for EachOption in OptionToColumn:
        ColumnNumber = ReverseNumbering(OptionToColumn[EachOption], len(OptionToColumn))
        CurrentCell = WorkSheet.cell(row= TopRow, column=ColumnNumber + 1)
        CurrentCell.value = EachOption

    CurrentCell = WorkSheet.cell(row= TopRow, column= len(OptionToColumn ) + 2)
    CurrentCell.value = 'Success'



def AppendConfigurationToWorkbook(ConfigInstance, Row, CurWorksheet, Success):
    Value = 'Value'

    CurrentCell = CurWorksheet.cell(row= Row, column= 1)
    CurrentCell.value = 'Test Number: ' + str(Row)

    for EachOption in OptionToColumn:
        ColumnNumber = ReverseNumbering(OptionToColumn[EachOption], len(OptionToColumn))
        CurrentCell = CurWorksheet.cell(row= Row, column= ColumnNumber + 1)
        CurrentCell.value = ConfigInstance.get(EachOption, Value)


    CurrentCell = CurWorksheet.cell(row= Row, column= len(OptionToColumn) + 2)
    CurrentCell.value = str(Success)


def BitToBool(BitString):
    return BitString == '1'

##
# 
#     Will check if a section has a requirement. If so, it checks if that requirement is satified.
#     Will only return false if the requirement is not satisfied.
#     @param ConfigInstance
#     @param SectionName
#     @return
#     
def CheckForRequired(ConfigInstance, SectionName):
    Required = 'Required'
    Value = 'Value'
    Passes = True
    assert isinstance(ConfigInstance, ConfigParser.SafeConfigParser)


    if ConfigInstance.has_option(SectionName, Required):
        Requirement = ConfigInstance.get(SectionName, Required)
        print('Section "' + SectionName + '" requires ' + Requirement)

        RequiredSection, RequiredValue = Requirement.split(' = ')
        if not ConfigInstance.get(RequiredSection, Value) == RequiredValue:
            print('Requirement not met. ' + RequiredSection + ' = ' + ConfigInstance.get(RequiredSection, Value))
            Passes = False

    return Passes

def ValidateConfiguration(ConfigInstance):

    Valid = True

    for Section in ConfigInstance.sections():
        if ConfigInstance.has_option(Section, 'SectionHeader'):
            continue

        try:
            if not ConfigInstance.getboolean(Section, 'Value'):
                continue
        except ValueError:
            pass

        if not CheckForRequired(ConfigInstance, Section):
            Valid = False
            break

    return Valid

def RunConfiguration(ConfigInstance):
    assert isinstance(ConfigInstance, ConfigParser.SafeConfigParser)
    CallString = 'python BuildIFWI.py '

    Flag = 'Flag'
    Value = 'Value'

    Success = True

    for Section in ConfigInstance.sections():
        if ConfigInstance.has_option(Section, 'SectionHeader'):
            continue
        try:
            if ConfigInstance.getboolean(Section, Value):
                CallString += ConfigInstance.get(Section, Flag) + ' '
        except ValueError:
            FlagString = ConfigInstance.get(Section, Flag) + '=' + ConfigInstance.get(Section, Value)
            CallString += FlagString + ' '

    print(CallString)
    error = subprocess.call(CallString)
    if error != 0:
        Success = False

    return Success


def MapBinaryToConfiguration(Binary, ConfigInstance):
    assert isinstance(ConfigInstance, ConfigParser.SafeConfigParser)
    Value = 'Value'

    for EachArg in OptionToColumn:
        ConfigInstance.set(EachArg, Value, bin(Binary)[len(bin(Binary)) - OptionToColumn[EachArg]])

    for EachKey in NonBools:
        BitValue = bin(Binary)[len(bin(Binary)) - OptionToColumn[EachKey]]
        StringValue = NonBools[EachKey][BitValue]
        ConfigInstance.set(EachKey, Value, StringValue)

    return ConfigInstance


def TestConfigurations(ConfigInstance):

    NumValidConfigurations = 0

    ExcelConfigFile = Workbook()
    ExcelDestination = 'ExampleExcel.xlsx'
    # CurrentWorksheet = ExcelConfigFile.active

    PlatformTypes = ['BXTM', 'APLK', 'APLI']
    for EachPlatform in PlatformTypes:
        TestNum = 1

        CurrentWorksheet = ExcelConfigFile.create_sheet()
        CurrentWorksheet.title = EachPlatform
        InitializeWorkSheet(CurrentWorksheet)


        ConfigNumber = BaseBinary
        UpperLimit = BaseBinary + (BaseBinary - 1)
        ConfigInstance.set('PlatformType', 'Value', EachPlatform)
        while ConfigNumber <= UpperLimit:
            ConfigInstance = MapBinaryToConfiguration(ConfigNumber, ConfigInstance)
            if ValidateConfiguration(ConfigInstance):
                NumValidConfigurations += 1
                TestNum += 1
                Success = RunConfiguration(ConfigInstance)
                AppendConfigurationToWorkbook(ConfigInstance, TestNum, CurrentWorksheet, Success)
                print(NumValidConfigurations)
                ExcelConfigFile.save(filename= ExcelDestination)
            ConfigNumber += 1

    print NumValidConfigurations
    ExcelConfigFile.save(filename= ExcelDestination)



if __name__ == '__main__':
    Config = ConfigParser.SafeConfigParser()
    Config.read('IfwiConfiguration.conf')
    TestConfigurations(Config)
